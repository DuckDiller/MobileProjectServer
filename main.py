import http.server
import openpyxl
import re
import json


class HttpProcessor(http.server.BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.send_header('content-type', 'text/html')
        self.end_headers()
        file = open('json.txt','r')
        output = file.read()
        file.close()
        b = bytes(output, 'utf-16')
        self.wfile.write(b)


class Group:
    def __init__(self, name, row, column):
        self.name = name
        self.row = row
        self.column = column
        self.subgroups = []
        self.subgroup_names = []
        self.cell_size = 0
        self.subjects = []

    def add_subgroup(self, subgroup):
        self.subgroups.append(subgroup)

    def add_subgroup_name(self, name):
        self.subgroup_names.append(name)

    def add_cell_size(self, size):
        self.cell_size = size

    def add_subject(self, subj):
        self.subjects.append(subj)

    # def toJSON(self):
    #     return json.dumps(self, default=lambda o: o.__dict__,
    #                       sort_keys=True, indent=4)
    # def obj_dict(obj):
    #     return obj.__dict__


class SubGroup:
    def __init__(self, name, row, column, size):
        self.name = name
        self.row = row
        self.column = column
        self.cell_size = size

    def toJSON(self):
        return json.dumps(self, default=lambda o: o.__dict__,
                          sort_keys=True, indent=4)


class Subject:
    def __init__(self, name, time, day_of_week):
        self.name = name
        self.time = time
        self.day_of_week = day_of_week
        self.subgroup_name = None

    def add_subgroup_name(self, name):
        self.subgroup_name = name

    def toJSON(self):
        return json.dumps(self, default=lambda o: o.__dict__,
                          sort_keys=True, indent=4)


def parse(sheet_name):
    book = openpyxl.load_workbook(
        filename=r'C:\Users\User\Downloads\Raspisanie-2-polugodie-2022-2023.-Filologi.xlsx')
    # print(book.sheetnames)
    sheet = book[sheet_name]
    row = 0
    column = 0
    groups = []
    for i in range(1, 10):
        if sheet[i][0].value == 'день/дата':
            print('A' + str(i))
            row = i
    column += 2  # column c
    print('row',row,'column',column)
    print(sheet[row][column])
    if sheet[row][column].value == 'УЧЕБНЫЕ ГРУППЫ':
        print('учебные группы')
        row += 1  # C5
        print(sheet[row][column].value)  # first group
        group_names = [sheet[row][column].value]
        groups.append(Group(name=sheet[row][column].value, row=row, column=column))
        print(group_names)
        out_of_groups = False
        column_upper = 1
        count_to_stop = 0
        print(sheet[row][column].value, row, column)
        while not out_of_groups:
            try:
                if sheet[row][column + column_upper].value is not None:
                    if sheet[row][column + column_upper].value not in group_names:
                        group_names.append(sheet[row][column + column_upper].value)
                        print('строка 58', sheet[row][column + column_upper].value, 'row', row, column + column_upper)
                        newgroup = Group(name=sheet[row][column + column_upper].value, row=row,
                                         column=column + column_upper)
                        groups.append(newgroup)
                else:
                    count_to_stop += 1
            except IndexError:
                print('tuple index out of range!')
                count_to_stop += 1
            if count_to_stop == 10:
                out_of_groups = True
            column_upper += 1
        print(group_names)
        print(groups[0].name, 'класс')
        column = 0
        row -= 1
        is_sub_groups = False
        print(str(chr(65 + column)) + str(row))  # A4
        for ranges in sheet.merged_cells.ranges:  # Проверяем на наличие подгрупп
            if str(chr(65 + column)) + str(row) in ranges:
                print(str(ranges))
                numbers = re.findall(r'\d+', str(ranges))  # Отделяет числа в отдельный массив
                print(numbers)
                if int(numbers[1]) - int(numbers[0]) > 1:
                    is_sub_groups = True
                    print(is_sub_groups)
        if is_sub_groups:  # если есть подгруппы
            column += 2
            row += 1
            print(sheet[row][column])  # C5
            for i in range(len(group_names)):
                for ranges in sheet.merged_cells.ranges:  # находим размер гориз. ячейки группы
                    if str(chr(65 + column)) + str(row) in ranges:
                        print(str(ranges))
                        columns = re.sub(r'[^A-z]', '', str(ranges))
                        print(columns)
                        cell_size = ord(columns[1]) - ord(columns[0]) + 1
                        print('размер ячейки', cell_size)
                        groups[i].add_cell_size(cell_size)
                row += 1  # C6
                size_of_subgroups = 0
                while size_of_subgroups < cell_size:
                    subgroup_cell_size = 1
                    for ranges in sheet.merged_cells.ranges:  # находим размер гориз. ячейки подгруппы
                        if str(chr(65 + column)) + str(row) in ranges:
                            print(str(ranges))
                            columns = re.sub(r'[^A-z]', '', str(ranges))
                            print(columns)
                            subgroup_cell_size = ord(columns[1]) - ord(columns[0]) + 1
                    print('размер ячейки подгруппы', subgroup_cell_size)
                    groups[i].add_subgroup(
                        SubGroup(name=sheet[row][column].value, row=row, column=column, size=subgroup_cell_size))
                    groups[i].add_subgroup_name(sheet[row][column].value)
                    size_of_subgroups += subgroup_cell_size
                    column += subgroup_cell_size
                row -= 1
            for i in groups:
                print(i.name, 'size', i.cell_size)
                for j in range(len(i.subgroups)):
                    print('\t', i.subgroups[j].name, i.subgroups[j].cell_size)
            size_of_all_groups = 0
            for i in groups:
                size_of_all_groups += i.cell_size
            for i in groups:
                column_adderator = 0
                for sub in range(len(i.subgroups)):
                    start_row = i.row + 1
                    column = i.column + column_adderator
                    for j in range(65):
                        if sheet[start_row][0].value is not None:
                            day_of_week = sheet[start_row][0].value
                        if sheet[start_row][1].value is not None:
                            time = sheet[start_row][1].value
                        if sheet[start_row][column].value is not None \
                                and sheet[start_row][column].value != i.name \
                                and sheet[start_row][column].value not in i.subgroup_names \
                                and sheet[start_row][column].value != 'УЧЕБНЫЕ ГРУППЫ':
                            subject_cell_size = 1
                            for ranges in sheet.merged_cells.ranges:  # находим размер гориз. ячейки текущего предмета
                                if str(chr(65 + column)) + str(start_row) in ranges:
                                    columns = re.sub(r'[^A-z]', '', str(ranges))
                                    subject_cell_size = ord(columns[1]) - ord(columns[0]) + 1
                            if subject_cell_size == size_of_all_groups:
                                for k in groups:
                                    k.add_subject(Subject(name=sheet[start_row][column].value, time=time,
                                                          day_of_week=day_of_week))
                            elif subject_cell_size == i.cell_size:
                                i.add_subject(
                                    Subject(name=sheet[start_row][column].value, time=time,
                                            day_of_week=day_of_week))
                            elif subject_cell_size == i.cell_size / len(i.subgroups):
                                subj = Subject(name=sheet[start_row][column].value, time=time,
                                               day_of_week=day_of_week)
                                subj.add_subgroup_name(i.subgroup_names[sub])
                                i.add_subject(subj)
                            elif subject_cell_size == i.cell_size * 2:
                                i.add_subject(
                                    Subject(name=sheet[start_row][column].value, time=time,
                                            day_of_week=day_of_week))
                                for search in range(len(groups)):
                                    if groups[search].name == i.name:
                                        groups[search + 1].add_subject(
                                            Subject(name=sheet[start_row][column].value, time=time,
                                                    day_of_week=day_of_week))
                        start_row += 1
                    column_adderator += i.subgroups[sub].cell_size
        else:  # если нет подгрупп
            column += 2
            row += 1
            print(sheet[row][column])  # C5
            for i in range(len(group_names)):
                cell_size = 1
                for ranges in sheet.merged_cells.ranges:  # находим размер гориз. ячейки группы
                    if str(chr(65 + column)) + str(row) in ranges:
                        print(str(ranges))
                        columns = re.sub(r'[^A-z]', '', str(ranges))
                        print(columns)
                        cell_size = ord(columns[1]) - ord(columns[0]) + 1
                        print('размер ячейки', cell_size)
                groups[i].add_cell_size(cell_size)
                column += cell_size
            size_of_all_groups = 0
            for i in groups:
                size_of_all_groups += i.cell_size
            for i in groups:
                start_row = i.row + 1
                column = i.column
                for j in range(65):
                    if sheet[start_row][0].value is not None:
                        day_of_week = sheet[start_row][0].value
                    if sheet[start_row][1].value is not None:
                        time = sheet[start_row][1].value
                    if sheet[start_row][column].value is not None \
                            and sheet[start_row][column].value != i.name \
                            and sheet[start_row][column].value != 'УЧЕБНЫЕ ГРУППЫ':
                        subject_cell_size = 1
                        for ranges in sheet.merged_cells.ranges:  # находим размер гориз. ячейки текущего предмета
                            if str(chr(65 + column)) + str(start_row) in ranges:
                                columns = re.sub(r'[^A-z]', '', str(ranges))
                                subject_cell_size = ord(columns[1]) - ord(columns[0]) + 1
                        if subject_cell_size == size_of_all_groups:
                            for k in groups:
                                k.add_subject(Subject(name=sheet[start_row][column].value, time=time,
                                                      day_of_week=day_of_week))
                        elif subject_cell_size == i.cell_size:
                            i.add_subject(
                                Subject(name=sheet[start_row][column].value, time=time, day_of_week=day_of_week))
                        elif subject_cell_size == i.cell_size * 2:
                            i.add_subject(
                                Subject(name=sheet[start_row][column].value, time=time, day_of_week=day_of_week))
                            for search in range(len(groups)):
                                if groups[search].name == i.name:
                                    groups[search + 1].add_subject(
                                        Subject(name=sheet[start_row][column].value, time=time,
                                                day_of_week=day_of_week))
                    start_row += 1
    return groups


def obj_dict(obj):
    return obj.__dict__


if __name__ == '__main__':
    book = openpyxl.load_workbook(
        filename=r'C:\Users\User\Downloads\Raspisanie-2-polugodie-2022-2023.-Filologi.xlsx')
    print(book.sheetnames)
    all_groups = []
    for sheet in book.sheetnames:
        groups = parse(sheet)
        for group in groups:
            all_groups.append(group)

    for i in all_groups:  # вывод
        print(i.name, 'size', i.cell_size)
        for j in range(len(i.subjects)):
            print('\t', i.subjects[j].subgroup_name, i.subjects[j].day_of_week, i.subjects[j].time,
                  i.subjects[j].name)

    x = json.dumps(all_groups, default=obj_dict, ensure_ascii=False)
    print(x)
    file = open('json.txt','w')
    file.write(x)
    file.close()
    serv = http.server.HTTPServer(("localhost", 4041), HttpProcessor)
    while True:
        serv.handle_request()